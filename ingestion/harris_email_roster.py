#!/usr/bin/env python3
"""
Harris Email Roster Importer

Reads CSV/XLSX roster files placed in a folder (default: ./email_rosters),
parses rows, and upserts:
 - Inserts each row into collection 'harris_email_roster' (with file reference)
 - Attempts to enrich existing 'harris_bond'/'harris_misfel'/'harris_nafiling' docs
   by SPN or case_number (and optionally name+dob) with any additional fields found.

Environment:
 - HARRIS_EMAIL_ROSTER_DIR: path to folder containing CSV/XLSX files (default: ./email_rosters)

Run via runner:
  python3 -m scripts.run_ingestion --source harris_email_roster
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
import re
import warnings


try:
    import xlrd  # type: ignore
except Exception:
    xlrd = None  # Optional: for legacy .xls files
import csv
import hashlib

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None  # CSV will still work without XLSX support


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


HEADER_MAP = {
    # normalized_field: [aliases]
    "spn": ["spn", "inmate_spn", "sid", "sid/spn", "spn id"],
    "case_number": ["case_number", "case no", "case", "cause", "cause_number"],
    "name": ["name", "full_name", "inmate_name", "defendant", "last, first", "last_first"],
    "last_name": ["last", "last_name", "surname"],
    "first_middle": ["first", "first_middle", "first name", "given"],
    "dob": ["dob", "date_of_birth", "birthdate", "d.o.b."],
    "offense": ["offense", "charge", "description"],
    "facility": ["facility", "location", "jail"],
    "status": ["status", "custody_status"],
    "booking_date": ["booking_date", "booked", "booked_date", "date booked"],
    "bond_amount": ["bond", "bond_amount", "bond amt"],
    # Emergency contact phone numbers
    "phone_nbr1": ["phone nbr1", "phone nbr 1", "phone1", "phone #1", "phone number 1"],
    "phone_nbr2": ["phone nbr2", "phone nbr 2", "phone2", "phone #2", "phone number 2"],
    "phone_nbr3": ["phone nbr3", "phone nbr 3", "phone3", "phone #3", "phone number 3"],
}


def _normalize_header(h: str) -> str:
    h = (h or "").strip().lower()
    for key, aliases in HEADER_MAP.items():
        if h == key or h in aliases:
            return key
    return h  # keep unknowns
    
def _debug_enabled() -> bool:
    return os.getenv("HARRIS_ROSTER_DEBUG", "0").strip() not in ("0", "false", "False", "")
    
def _is_document_map(sheet_name: str) -> bool:
    s = (sheet_name or "").strip().lower().replace(" ", "")
    return s == "documentmap"

def _is_allowed_roster_sheet(sheet_name: str) -> bool:
    # Allow any non-empty sheet except 'Document Map'
    s = (sheet_name or "").strip()
    if not s:
        return False
    return not _is_document_map(s)


def _parse_csv(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        headers_raw = next(reader, [])
        headers = [_normalize_header(h) for h in headers_raw]
        for r in reader:
            if not any(r):
                continue
            row = {headers[i]: (r[i].strip() if i < len(r) else None) for i in range(len(headers))}
            rows.append(row)
    return rows


## removed: old single-sheet _parse_xlsx that used wb.active
    
def _parse_xlsx(path: Path) -> List[Dict[str, Any]]:
    assert openpyxl is not None, "openpyxl not installed; cannot read .xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    if _debug_enabled():
        try:
            print(f"[roster.debug] xlsx file={path.name} sheetnames={wb.sheetnames}")
        except Exception:
            pass
    processed_any = False
    for name in wb.sheetnames:
        if _is_document_map(name) or not _is_allowed_roster_sheet(name):
            continue
        ws = wb[name]
        headers: List[str] = []
        first = True
        n_before = len(rows)
        for row in ws.iter_rows(values_only=True):
            if first:
                headers_raw = [str(c) if c is not None else "" for c in row or []]
                headers = [_normalize_header(h) for h in headers_raw]
                first = False
                continue
            vals = [str(c).strip() if c is not None else None for c in (row or [])]
            if not any(vals):
                continue
            rec = {headers[j]: (vals[j] if j < len(vals) else None) for j in range(len(headers))}
            rows.append(rec)
        processed_any = True
        if _debug_enabled():
            try:
                print(f"[roster.debug] xlsx processed sheet='{name}' added_rows={len(rows)-n_before}")
            except Exception:
                pass
    if _debug_enabled() and not processed_any:
        try:
            print(f"[roster.debug] xlsx file={path.name} no allowed sheets found (only 'Sheet1/2/…' are processed; 'Document Map' skipped)")
        except Exception:
            pass
    return rows


def _parse_xls(path: Path) -> List[Dict[str, Any]]:
    assert xlrd is not None, "xlrd not installed; cannot read .xls"
    # Suppress noisy OLE/xlrd warnings for legacy .xls files
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = xlrd.open_workbook(path, on_demand=True)
    rows: List[Dict[str, Any]] = []
    names = wb.sheet_names()
    if _debug_enabled():
        try:
            print(f"[roster.debug] xls  file={path.name} sheetnames={names}")
        except Exception:
            pass
    processed_any = False
    for name in names:
        if _is_document_map(name) or not _is_allowed_roster_sheet(name):
            continue
        sh = wb.sheet_by_name(name)
        headers: List[str] = []
        n_before = len(rows)
        for i in range(sh.nrows):
            row_vals = sh.row_values(i)
            if i == 0:
                headers_raw = [str(c) if c is not None else "" for c in row_vals]
                headers = [_normalize_header(h) for h in headers_raw]
                continue
            vals = [str(c).strip() if c is not None else None for c in row_vals]
            if not any(vals):
                continue
            rec = {headers[j]: (vals[j] if j < len(vals) else None) for j in range(len(headers))}
            rows.append(rec)
        processed_any = True
        if _debug_enabled():
            try:
                print(f"[roster.debug] xls  processed sheet='{name}' added_rows={len(rows)-n_before}")
            except Exception:
                pass
    if _debug_enabled() and not processed_any:
        try:
            print(f"[roster.debug] xls  file={path.name} no allowed sheets found (only 'Sheet1/2/…' are processed; 'Document Map' skipped)")
        except Exception:
            pass
    return rows


def _coalesce_name(row: Dict[str, Any]) -> Optional[str]:
    if row.get("name"):
        return str(row["name"]).strip()
    last = (row.get("last_name") or "").strip()
    first = (row.get("first_middle") or "").strip()
    if last or first:
        return ", ".join([x for x in [last, first] if x])
    return None


@dataclass
class HarrisEmailRosterImporter:
    db: Any

    def _target_dir(self) -> Path:
        d_raw = os.getenv("HARRIS_EMAIL_ROSTER_DIR", "email_rosters").strip()
        # Support Dropbox paths like "~/Dropbox/..." and env vars like "$HOME/Dropbox/..."
        d = os.path.expanduser(os.path.expandvars(d_raw))
        p = Path(d)
        if not p.is_absolute():
            # relative to repo root (scripts run from repo root)
            p = Path.cwd() / p
        p.mkdir(parents=True, exist_ok=True)
        return p

    def _list_files(self) -> List[Path]:
        base = self._target_dir()
        return sorted([*base.glob("**/*.csv"), *base.glob("**/*.xlsx"), *base.glob("**/*.xls")])

    def _file_sha256(self, path: Path) -> str:
        h = hashlib.sha256()
        with path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()

    def _parse_file(self, path: Path) -> List[Dict[str, Any]]:
        try:
            if path.suffix.lower() == ".csv":
                return _parse_csv(path)
            elif path.suffix.lower() == ".xlsx":
                return _parse_xlsx(path)
            elif path.suffix.lower() == ".xls":
                return _parse_xls(path)
        except Exception as e:
            print(f"[harris_email_roster] parse error {path.name}: {e}")
        return []

    def _ensure_indexes(self):
        # roster collection
        roster = self.db["harris_email_roster"]
        try:
            roster.create_index([("spn", 1)], background=True)
            roster.create_index([("case_number", 1)], background=True)
            roster.create_index([("name", 1)], background=True)
            roster.create_index([("dob", 1)], background=True)
            roster.create_index([("loaded_at", 1)], background=True)
        except Exception:
            pass
        # processed files ledger
        files = self.db["harris_email_roster_files"]
        try:
            files.create_index([("sha256", 1)], unique=True, background=True)
            files.create_index([("processed_at", 1)], background=True)
            files.create_index([("file_name", 1)], background=True)
        except Exception:
            pass
        # harris_* collections (may already exist)
        for name in ("harris_bond", "harris_misfel", "harris_nafiling"):
            col = self.db[name]
            try:
                col.create_index([("spn", 1)], background=True)
                col.create_index([("case_number", 1)], background=True)
                col.create_index([("updated_at", 1)], background=True)
                col.create_index([("history_hashes", 1)], background=True)
            except Exception:
                pass

    def _enrich_existing(self, row: Dict[str, Any], file_sha: Optional[str]) -> int:
        """Try to update an existing harris_* doc with additional fields from the row.
        Returns number of documents modified across all target collections.
        """
        spn = (row.get("spn") or "").strip()
        case = (row.get("case_number") or "").strip()
        name = _coalesce_name(row)
        dob = (row.get("dob") or "").strip()

        # Build a $set payload with only non-empty fields
        set_fields = {}
        for k in ("facility", "status", "offense", "booking_date", "bond_amount"):
            v = row.get(k)
            if v not in (None, ""):
                set_fields[k] = v
        if name and not set_fields.get("name"):
            set_fields["name"] = name

        if not set_fields:
            return 0

        mod_total = 0
        filt_or = []
        if spn:
            filt_or.append({"spn": spn})
        if case:
            filt_or.append({"case_number": case})
        # optional: name+dob fallback
        if name and dob:
            filt_or.append({"name": name, "dob": dob})
        if not filt_or:
            return 0

        # Always set fields
        for col_name in ("harris_bond", "harris_misfel", "harris_nafiling"):
            res_set = self.db[col_name].update_many({"$or": filt_or}, {"$set": {**set_fields, "updated_at": _now_iso()}})
            mod_total += int(res_set.modified_count or 0)
            # Conditionally push history once per file hash
            hist_entry = {"ts": _now_iso(), "source": "email_roster", **set_fields}
            try:
                if file_sha:
                    self.db[col_name].update_many(
                        {"$and": [{"$or": filt_or}, {"history_hashes": {"$ne": file_sha}}]},
                        {"$push": {"history": hist_entry}, "$addToSet": {"history_hashes": file_sha}},
                    )
                else:
                    self.db[col_name].update_many({"$or": filt_or}, {"$push": {"history": hist_entry}})
            except Exception:
                pass
        return mod_total

    def _upsert_simple_harris_phones(self, row: Dict[str, Any]) -> int:
        """Upsert emergency contact phone numbers into simple_harris.
        Matches by SPN primarily; falls back to name+dob if present.
        Returns modified count.
        """
        def _get_phone_variants(row: Dict[str, Any], idx: int) -> str:
            # try normalized first
            norm_key = f"phone_nbr{idx}"
            val = row.get(norm_key)
            if isinstance(val, str) and val.strip():
                return val.strip()
            # common variants observed in roster sheets (spaces, symbols, NBSP)
            candidates = [
                f"phone nbr{idx}",
                f"phone nbr {idx}",
                f"phone#{idx}",
                f"phone # {idx}",
                f"phone #"+str(idx),
                f"phone number {idx}",
                f"phone{idx}",
            ]
            for k in candidates:
                v = row.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip()
            # last resort: search keys loosely
            try:
                for k, v in row.items():
                    if not isinstance(k, str) or not isinstance(v, str):
                        continue
                    ks = k.lower().replace("\u00a0", " ").strip()
                    if ks.startswith("phone") and any(t in ks for t in ("nbr", "number", "#")) and ks.endswith(str(idx)):
                        if v.strip():
                            return v.strip()
            except Exception:
                pass
            return ""

        ph1 = _get_phone_variants(row, 1)
        ph2 = _get_phone_variants(row, 2)
        ph3 = _get_phone_variants(row, 3)
        # If no phones provided, nothing to do
        if not any([ph1, ph2, ph3]):
            return 0

        spn = (row.get("spn") or "").strip()
        case = (row.get("case_number") or "").strip()
        name = _coalesce_name(row)
        dob = (row.get("dob") or "").strip()

        filt_or = []
        if spn:
            filt_or.append({"spn": spn})
        # As a secondary, allow match by case_number if present (less strict)
        if case:
            filt_or.append({"case_number": case})
        if name and dob:
            filt_or.append({"full_name": name, "dob": dob})
        if not filt_or:
            return 0

        set_fields = {"phones_updated_at": _now_iso(), "phones_source": "harris_email_roster"}
        if ph1:
            set_fields["phone_nbr1"] = ph1
        if ph2:
            set_fields["phone_nbr2"] = ph2
        if ph3:
            set_fields["phone_nbr3"] = ph3

        res = self.db["simple_harris"].update_many({"$or": filt_or}, {"$set": set_fields})
        return int(res.modified_count or 0)

    def run(self) -> Dict[str, Any]:
        self._ensure_indexes()
        files = self._list_files()
        try:
            print(f"[harris_email_roster] target_dir={self._target_dir()} files_found={len(files)}")
        except Exception:
            pass
        force_reprocess = os.getenv("HARRIS_ROSTER_FORCE_REPROCESS", "0").strip() not in ("0", "false", "False", "")
        inserted_roster = 0
        updated_existing = 0
        skipped_files: List[str] = []
        skipped_duplicates = 0
        total_rows = 0
        roster_col = self.db["harris_email_roster"]
        ledger_col = self.db["harris_email_roster_files"]

        for f in files:
            # Skip exact duplicate files by content hash
            try:
                sha = self._file_sha256(f)
            except Exception as e:
                print(f"[harris_email_roster] hash error {f.name}: {e}")
                sha = None

            if sha is not None:
                existing = ledger_col.find_one({"sha256": sha})
                if existing and not force_reprocess:
                    # Skip only if we previously processed and got rows > 0
                    prev_rows = int(existing.get("rows", 0) or 0)
                    if prev_rows > 0:
                        skipped_duplicates += 1
                        skipped_files.append(f.name)
                        continue

            rows = self._parse_file(f)
            if not rows:
                skipped_files.append(f.name)
                continue
            file_inserted = 0
            file_updated = 0
            for r in rows:
                total_rows += 1
                # normalize keys
                r = {k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
                r.setdefault("spn", (r.get("spn") or "").strip())
                r.setdefault("case_number", (r.get("case_number") or "").strip())
                r.setdefault("name", _coalesce_name(r))
                r.setdefault("source", "harris_email_roster")
                r.setdefault("loaded_at", _now_iso())
                r.setdefault("file_ref", str(f))

                # insert into roster collection (upsert by spn/case_number/name+dob)
                filt = None
                if r.get("spn") and r.get("case_number"):
                    filt = {"spn": r["spn"], "case_number": r["case_number"]}
                elif r.get("spn"):
                    filt = {"spn": r["spn"]}
                elif r.get("case_number"):
                    filt = {"case_number": r["case_number"]}
                elif r.get("name") and r.get("dob"):
                    filt = {"name": r["name"], "dob": r["dob"]}
                if filt is None:
                    # no reliable identifier; still store it with a random key
                    filt = {"file_ref": str(f), "row_index": total_rows}

                res = roster_col.update_one(filt, {"$set": r, "$setOnInsert": {"first_loaded_at": _now_iso()}}, upsert=True)
                if res.upserted_id is not None:
                    inserted_roster += 1
                    file_inserted += 1

                # enrich existing harris_* docs
                inc = self._enrich_existing(r, sha)
                updated_existing += inc
                file_updated += inc

                # update simple_harris phones if present
                try:
                    ph_mod = self._upsert_simple_harris_phones(r)
                    file_updated += ph_mod
                except Exception as e:
                    if _debug_enabled():
                        print(f"[roster.debug] simple_harris phone update error: {e}")

            # Mark file as processed in ledger (after successful parse/process)
            try:
                stat = f.stat()
                ledger_doc = {
                    "sha256": sha or "",
                    "file_name": f.name,
                    "file_path": str(f),
                    "size": stat.st_size,
                    "mtime": stat.st_mtime,
                    "processed_at": _now_iso(),
                    "rows": len(rows),
                    "inserted_roster": file_inserted,
                    "updated_existing": file_updated,
                }
                if sha is not None:
                    # Upsert so we can correct prior entries that had rows==0
                    ledger_col.update_one({"sha256": sha}, {"$set": ledger_doc}, upsert=True)
            except Exception as e:
                print(f"[harris_email_roster] ledger write failed for {f.name}: {e}")

        return {
            "files": [str(p) for p in files],
            "total_rows": total_rows,
            "inserted_roster": inserted_roster,
            "updated_existing": updated_existing,
            "skipped_files": skipped_files,
            "skipped_duplicates": skipped_duplicates,
        }
